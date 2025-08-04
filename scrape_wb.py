import asyncio
import time
import functools
from typing import List, Optional

import httpx
from openpyxl import Workbook
from pydantic import BaseModel


MAIN_MENU_URL = "https://static-basket-01.wbbasket.ru/vol0/data/main-menu-by-ru-v2.json"
FILTER_API = "https://catalog.wb.ru/catalog/{shard}/v4/filters?appType=1&{query}&curr=rub&dest=-59202"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/125.0.0.0 Safari/537.36",
    "Accept": "*/*",
    "Origin": "https://www.wildberries.ru",
    "Referer": "https://www.wildberries.ru/"
}

def benchmark(func):
    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        start = time.perf_counter()
        result = await func(*args, **kwargs)
        elapsed = time.perf_counter() - start
        print(f"{func.__name__} completed in {elapsed:.2f} sec.")
        return result
    return wrapper

def safe_api_call(func):
    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        try:
            return await func(*args, **kwargs)
        except Exception as e:
            print(f"Error in {func.__name__}: {e}")
            return None
    return wrapper

class Category(BaseModel):
    id: int
    name: str
    shard: Optional[str] = None
    query: Optional[str] = None
    url: Optional[str] = None
    childs: Optional[List['Category']] = None

    model_config = {
        "from_attributes": True,
        "arbitrary_types_allowed": True
    }

Category.model_rebuild()

class FilterItem(BaseModel):
    id: int
    name: str
    depth: int = 99
    parent: Optional[int] = None

class FilterGroup(BaseModel):
    name: Optional[str]
    items: Optional[List[FilterItem]]

class Filters(BaseModel):
    filters: Optional[List[FilterGroup]]

class ExcelSaver:
    def __init__(self):
        self.workbook = Workbook()
        self.sheets = {}

    def get_sheet(self, name: str):
        if name not in self.sheets:
            ws = self.workbook.create_sheet(title=name[:31])
            ws.append(["ID", "Name", "Depth", "Parent"])
            self.sheets[name] = ws
        return self.sheets[name]

    def write(self, sheet_name: str, item_id: int, name: str, depth: int, parent: Optional[int]):
        sheet = self.get_sheet(sheet_name)
        sheet.append([item_id, name, depth, parent or 0])

    def save(self, filename="wb_cat.xlsx"):
        if "Sheet" in self.workbook.sheetnames and len(self.workbook.sheetnames) > 1:
            del self.workbook["Sheet"]
        self.workbook.save(filename)

class WildberriesScraper:
    def __init__(self):
        self.excel = ExcelSaver()
        self.client = httpx.AsyncClient(headers=HEADERS, timeout=10)
        self.semaphore = asyncio.Semaphore(30)

    @benchmark
    async def scrape(self):
        catalog = await self._fetch_main_menu()
        if catalog:
            await self._process_all(catalog)
        self.excel.save()
        await self.client.aclose()

    @safe_api_call
    async def _fetch_main_menu(self) -> Optional[List[Category]]:
        resp = await self.client.get(MAIN_MENU_URL)
        data = resp.json()
        return [Category(**item) for item in data]

    async def _process_all(self, categories: List[Category]):
        tasks = [self._recurse(cat, 1, None, cat.name) for cat in categories]
        await asyncio.gather(*tasks)

    async def _recurse(self, cat: Category, depth: int, parent_id: Optional[int], sheet_name: str):
        self.excel.write(sheet_name, cat.id, cat.name, depth, parent_id)

        if cat.childs:
            await asyncio.gather(*[
                self._recurse(sub, depth + 1, cat.id, sheet_name)
                for sub in cat.childs
            ])
        elif cat.shard and cat.query:
            await self._fetch_items(cat, sheet_name)

    @safe_api_call
    async def _fetch_items(self, cat: Category, sheet_name: str):
        async with self.semaphore:
            url = FILTER_API.format(shard=cat.shard, query=cat.query)
            resp = await self.client.get(url)
            data = resp.json().get("data", {})
            raw_filters = data.get("filters", [])
            for group in raw_filters:
                if group.get("name") == "Категория" and isinstance(group.get("items"), list):
                    for item in group["items"]:
                        self.excel.write(
                            sheet_name,
                            item.get("id", 0),
                            item.get("name", "Unnamed"),
                            99,
                            cat.id
                        )


def main():
    asyncio.run(WildberriesScraper().scrape())

if __name__ == "__main__":
    main()
